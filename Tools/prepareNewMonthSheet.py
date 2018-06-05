from os import getcwd, path
from os.path import dirname
from shutil import copy2
from datetime import date, datetime
from calendar import Calendar
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.styles.borders import Border, Side

# functions
def clearCell(cell, fill, border):
    """
    Clears values from cells, removes/applies fill and removes borders (optional).
    :param cell: Cell object; cell to work on
    """
    cell.value = None
    if fill: cell.fill = PatternFill(start_color=fill, end_color=fill, fill_type="solid")
    else: cell.fill = PatternFill(fill_type=None)
    if border: cell.border = Border(left=Side(style=None), right=Side(style=None), top=Side(style=None), bottom=Side(style=None))

def formatCell(cell, border_top, border_right, border_bottom, border_left, fill):
    """
    Applies border to cell.
    :param cell: Cell object; cell to work on
    """
    cell.border = Border(top=Side(style=border_top), right=Side(style=border_right), bottom=Side(style=border_bottom), left=Side(style=border_left))
    if fill: cell.fill = PatternFill(start_color=fill, end_color=fill, fill_type="solid")
    else: cell.fill = PatternFill(fill_type=None)

def applyBorders(spreadsheet, top_border, bottom_border, top_left_border, top_right_border, bottom_left_border, bottom_right_border):
    """
    (Re)applies borders to cells in a spreadsheet.
    :param sheet: openpyxl object; active XLS[X] spreadsheet to work on
    """
    counter = 1
    for i in range(1, len(spreadsheet["A"]), 2):
        if spreadsheet["A"][i].value:
            counter += 1
            for column in ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O", "P", "Q"]:
                spreadsheet[column + str(i + 1)].border = top_border
                spreadsheet[column + str(i + 2)].border = bottom_border
            if month in [1, 3, 5, 7, 8, 10, 12]:
                for j in range(1, 32):
                    sheet[days[j][0] + str(j + 1)].border = top_left_border
                    sheet[days[j][1] + str(j + 1)].border = top_right_border
                    sheet[days[j][0] + str(j + 2)].border = bottom_left_border
                    sheet[days[j][1] + str(j + 2)].border = bottom_right_border
        else:
            counter += 1
            for column in ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O", "P", "Q", "R"]:
                spreadsheet[column + str(i + 1)].border = top_border
                spreadsheet[column + str(i + 2)].border = bottom_border
            break
    sheet["F" + str(counter * 2)].border = top_border
    sheet["F" + str(counter * 2 + 1)].border = bottom_border
    sheet["G" + str(counter * 2)].border = top_border
    sheet["G" + str(counter * 2 + 1)].border = bottom_border
    sheet["H" + str(counter * 2)].border = top_border
    sheet["H" + str(counter * 2 + 1)].border = bottom_border

# creates path & backup
path = getcwd().replace("\\", "\\\\") + "\\\\"
parent_dir = dirname(getcwd()).replace("\\", "\\\\") + "\\\\"
copy2(parent_dir + "Poslovni-pregled.xlsx", parent_dir + "Poslovni-pregled_new_month_backup.xlsx")
print("Backup file created: {}Poslovni-pregled_new_month_backup.xlsx".format(parent_dir.replace("\\\\", "\\")))
# date variables
months = { 1: "JAN",
           2: "FEB",
           3: "MAR",
           4: "APR",
           5: "MAJ",
           6: "JUN",
           7: "JUL",
           8: "AVG",
           9: "SEP",
          10: "OKT",
          11: "NOV",
          12: "DEC"}
days =   { 1: ("S", "T"),
           2: ("U", "V"),
           3: ("W", "X"),
           4: ("Y", "Z"),
           5: ("AA", "AB"),
           6: ("AC", "AD"),
           7: ("AE", "AF"),
           8: ("AG", "AH"),
           9: ("AI", "AJ"),
          10: ("AK", "AL"),
          11: ("AM", "AN"),
          12: ("AO", "AP"),
          13: ("AQ", "AR"),
          14: ("AS", "AT"),
          15: ("AU", "AV"),
          16: ("AW", "AX"),
          17: ("AY", "AZ"),
          18: ("BA", "BB"),
          19: ("BC", "BD"),
          20: ("BE", "BF"),
          21: ("BG", "BH"),
          22: ("BI", "BJ"),
          23: ("BK", "BL"),
          24: ("BM", "BN"),
          25: ("BO", "BP"),
          26: ("BQ", "BR"),
          27: ("BS", "BT"),
          28: ("BU", "BV"),
          29: ("BW", "BX"),
          30: ("BY", "BZ"),
          31: ("CA", "CB")}
today = date.today()
year = str(today.year)[2:]
month = months[today.month]
day = today.day
days_in_current_month = max(max(Calendar().monthdayscalendar(today.year, today.month)))
prev_month_year = today.year
prev_month_month = today.month - 1
if prev_month_month == 0: # decrements year and sets previous month to 12 (December)
    prev_month_year = prev_month_year - 1
    prev_month_month = 12
if prev_month_month < 10: # stringifies previous year and month, adds 0 in front of month if month == single digit
    prev_month_year = str(prev_month_year)
    prev_month_month = "0" + str(prev_month_month)
else:
    prev_month_year = str(prev_month_year)
    prev_month_month = str(prev_month_month)
# borders
top_border                = Border(top=Side(style='thin'))
bottom_border             = Border(bottom=Side(style='thin'))
left_border               = Border(left=Side(style='thin'))
right_border              = Border(right=Side(style='thin'))
top_border_sheet          = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'))
bottom_border_sheet       = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             bottom=Side(style='thin'))
top_left_border_sheet     = Border(left=Side(style='thin'),
                             top=Side(style='thin'))
top_right_border_sheet    = Border(right=Side(style='thin'),
                             top=Side(style='thin'))
bottom_left_border_sheet  = Border(left=Side(style='thin'),
                             bottom=Side(style='thin'))
bottom_right_border_sheet = Border(right=Side(style='thin'),
                             bottom=Side(style='thin'))
# creates new sheet for current month if no such sheet present
wb = load_workbook(parent_dir.replace("\\\\", "/") + "Poslovni-pregled.xlsx", data_only = True)
try:
    sheet = wb[month + " " + year]
except KeyError: # create sheet for new month if not existing
    decrement_month = 1
    decrement_year = 1
    while True: # cycles back by month until last valid sheet is found
        last_found_month = today.month - decrement_month
        last_found_year = today.year
        if last_found_month == 0:
            last_found_month = 12
            last_found_year = last_found_year - decrement_year
        try:
            last_found_sheet = wb[months[last_found_month] + " " + str(last_found_year)[2:]]
            sheet = wb.copy_worksheet(last_found_sheet)
            sheet.title = month + " " + year
            break
        except KeyError:
            decrement_month += 1
            decrement_year = decrement_month // 12 + 1
    counter = 1
    for i in range(1, len(sheet["A"]), 2):
        if sheet["A"][i].value: counter += 1
    for i in range(0, counter * 2): # clears values, fill and border from day and header cells from 29th to 31st
        for ii in range(29, 32):
            clearCell(sheet[days[ii][0] + str(i + 1)], None, 1)
            clearCell(sheet[days[ii][1] + str(i + 1)], None, 1)
    for i in range(1, counter * 2): # clears values and fill from day cells from 1st to 28th
        for ii in range(1, 29):
            clearCell(sheet[days[ii][0] + str(i + 1)], None, 0)
            clearCell(sheet[days[ii][1] + str(i + 1)], None, 0)
        for column in ("I", "J", "K", "L", "M", "N", "O", "P", "Q"): # clears values from data cells and applies grey fill
            clearCell(sheet[column + str(i + 1)], 'F2F2F2', 0)
        for column in ("G", "H"): # clears values from G and F columns if present
            if i % 2 != 0 and sheet[column + str(i + 1)].value != "-": clearCell(sheet[column + str(i + 1)], 'F2F2F2', 0)
    clearCell(sheet["F" + str((counter + 1) * 2)], 'F2F2F2', 0) # clears budget value
    clearCell(sheet["G" + str((counter + 1) * 2)], 'F2F2F2', 0) # clears total cost value
    clearCell(sheet["H" + str((counter + 1) * 2)], 'F2F2F2', 0) # clears total cost % value
    clearCell(sheet["A" + str((counter + 2) * 2 + 1)], None, 0) # clears last updated info
    for i in range(1, counter * 2, 2): # applies border around day cells from 29th to 31st
        for ii in range(29, days_in_current_month + 1):
            formatCell(sheet[days[ii][0] + str(i + 1)], 'thin', None, None, 'thin', None)
            formatCell(sheet[days[ii][1] + str(i + 1)], 'thin', 'thin', None, None, None)
            formatCell(sheet[days[ii][0] + str(i + 2)], None, None, 'thin', 'thin', None)
            formatCell(sheet[days[ii][1] + str(i + 2)], None, 'thin', 'thin', None, None)
    for i in range(29, days_in_current_month + 1): # applies border, shading and numbers to header cells from 29th to 31st
        formatCell(sheet[days[i][0] + "1"], 'thin', None, 'thin', 'thin', 'D9D9D9')
        formatCell(sheet[days[i][1] + "1"], 'thin', 'thin', 'thin', None, 'D9D9D9')
        sheet[days[i][0] + "1"] = i
    applyBorders(sheet, top_border_sheet, bottom_border_sheet, top_left_border_sheet, top_right_border_sheet, bottom_left_border_sheet, bottom_right_border_sheet)
    applyBorders(last_found_sheet, top_border_sheet, bottom_border_sheet, top_left_border_sheet, top_right_border_sheet,
                 bottom_left_border_sheet, bottom_right_border_sheet)
    sheet.freeze_panes = sheet["S2"]
while True:
    try:
        wb.save(parent_dir + "Poslovni-pregled.xlsx")
        break
    except PermissionError: input("\nPlease close Poslovni-pregled.xlsx and press any key. ")
print("\nNew month sheet added to {}Poslovni-pregled.xlsx".format(parent_dir.replace("\\\\", "\\")))