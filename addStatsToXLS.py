from os import getcwd
from shutil import copy2
import glob
import json
from datetime import date, datetime
from calendar import Calendar
from csv import reader
import requests
from bs4 import BeautifulSoup as bs
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.styles import PatternFill
from openpyxl.styles.borders import Border, Side
from re import sub

# functions
def mergePausedFiles(cwd_path):
    """
    Merges multiple paused GAdW campaign data (if any) and returns a list of JSON/dict objects.
    :param cwd_path: str; CWD path with GAdW_JSON_paused*.txt files to be processed
    :return: list; list of JSON/dict objects with paused GAdW campaign data
    """
    filepath = cwd_path + "GAdW_JSON_paused*.txt"
    txt = glob.glob(filepath)
    merged = list()
    for textfile in txt:
        try: paused = json.loads(sub(r"\s{2,}", " ", sub(r"\n", " ", open(textfile, encoding="UTF-8").read()))) # imports possible data on automatically paused campaigns
        except ValueError: paused = json.loads(sub(r"\s{2,}", " ", sub(r"\n", " ", open("GAdW_JSON_paused.txt", encoding="UTF-8-SIG").read()))) # handles possible ValueError due to manual TXT file editing
        except FileNotFoundError: paused = None
        if paused: merged.append(paused)
    return merged

def pauseToEndOfMonth(spreadsheet, start_day, days_map, days_in_month, fill, platform):
    """
    Marks cells as paused (applies fill) from input day to the end of month.
    :param spreadsheet: openpyxl object; active XLS[X] spreadsheet to work on
    :param start_day: int, start day to pause campaign from
    :param days_map: dict, mapping of day keys (int) to spreadsheet values (str)
    :param days_in_month: int; number of days in current month
    :param fill: PatternFill object; fill to apply to cells
    :param platform: str; platform to apply fill in, GAdW or TSmedia
    """
    assert platform in ("GAdW", "TSmedia")
    if platform == "GAdW": offset = 1
    else: offset = 2
    for d in range(start_day, days_in_month + 1):
        spreadsheet[days_map[d][0] + str(i + offset)].fill = fill
        spreadsheet[days_map[d][1] + str(i + offset)].fill = fill

def unpauseToEndOfMonth(spreadsheet, start_day, days_map, days_in_month, fill, platform):
    """
    Marks cells as not paused (removes fill) from input day to the end of month.
    :param spreadsheet: openpyxl object; active XLS[X] spreadsheet to work on
    :param start_day: int, start day to unpause campaign from
    :param days_map: dict, mapping of day keys (int) to spreadsheet values (str)
    :param days_in_month: int; number of days in current month
    :param fill: PatternFill object; fill to apply to cells
    :param platform: str; platform to remove fill in, GAdW or TSmedia
    """
    assert platform in ("GAdW", "TSmedia")
    if platform == "GAdW": offset = 1
    else: offset = 2
    for d in range(start_day, days_in_month + 1):
        spreadsheet[days_map[d][0] + str(i + offset)].fill = fill
        spreadsheet[days_map[d][1] + str(i + offset)].fill = fill

def findLastValue(spreadsheet, index, days_map, day_to_check):
    """
    Checks previous days recursively until the last entered value is found
    :param spreadsheet: openpyxl object; active XLS[X] spreadsheet to work on
    :param index: int; index of spreadsheet row to work on
    :param days_map: dict, mapping of day keys (int) to spreadsheet values (str)
    :param day_to_check: int; day prior to which to check for values
    :return: int; last value found or 0 if None
    """
    if day_to_check == 1 and not spreadsheet[days_map[day_to_check][0]][index].value: return 0
    if sheet[days[day_to_check][0]][index].value: return sheet[days[day_to_check][0]][index].value
    else: return findLastValue(spreadsheet, index, days_map, day_to_check - 1)

def performance(running, sheet, CPD, goal, day, days, days_in_current_month, index, platform=0, platforms=True):
    """
    Performance measures the pace of accruing clicks on GAdW and TSmedia networks:
        performance > 100% implies monthly goal will be met before the end of month
        performance == 100% implies monthly goal will be met at the end of month
        performance => 50% implies monthly goal will NOT be met at the end of month but campaign is still performing reasonably well
        performance < 50% implies monthly goal will NOT be met at the end of month and campaign is performing poorly
    :param running: tuple of size 2; platforms and current clicks, None if platform not present or no clicks value
    :param sheet: openpyxl object; active XLS[X] spreadsheet to work on
    :param CPD: float; clicks per day
    :param goal: int, monthly goal for platform
    :param day: int, current day in month
    :param days: dict, mapping of day keys (int) to spreadsheet values (str)
    :param days_in_current_month: int; number of days in current month
    :param index: int; index of spreadsheet row to work on
    :param platform: int; platform to work on, GAdW = 0, TSmedia = 1
    :param platforms: bool; True iff GAdW & TSmedia, False otherwise
    """

    def calculateAndFill(sheet, CPD, goal, day, days, days_in_current_month, index):
        """
        Helper function for performance(), calculates performance according to platforms and applies fill.
        :param sheet: openpyxl object; active XLS[X] spreadsheet to work on
        :param CPD: float; clicks per day
        :param goal: int, monthly goal for platform
        :param day: int, current day in month
        :param days: dict, mapping of day keys (int) to spreadsheet values (str)
        :param days_in_current_month: int; number of days in current month
        :param index: int; index of spreadsheet row to work on
        """
        if goal > 0:
            sheet[days[day][1] + str(index + 1)] = (CPD * days_in_current_month) / goal
            if sheet[days[day][1] + str(index + 1)].value > 1: sheet[days[day][1] + str(index + 1)].fill = dark_green
            elif sheet[days[day][1] + str(index + 1)].value == 1: sheet[days[day][1] + str(index + 1)].fill = light_green
            elif sheet[days[day][1] + str(index + 1)].value >= 0.75: sheet[days[day][1] + str(index + 1)].fill = blue_green
            elif sheet[days[day][1] + str(index + 1)].value >= 0.5: sheet[days[day][1] + str(index + 1)].fill = yellow
            elif sheet[days[day][1] + str(index + 1)].value >= 0.25: sheet[days[day][1] + str(index + 1)].fill = orange
            else: sheet[days[day][1] + str(index + 1)].fill = red
        else:
            sheet[days[day][1] + str(index + 1)] = "N/A"
            sheet[days[day][1] + str(index + 1)].alignment = Alignment(horizontal='right')
            sheet[days[day][1] + str(index + 1)].fill = dark_green
    if platform == 1: other = 0
    if platform == 0: other = 1
    if running[other][1]:  # checks if both platforms are running
        if platform == 1 and sheet[days[day][0]][index].value > sheet["L"][index].value: # if GAdW goal is surpassed, TSmedia performance is calculated with goal reduced
            possible_surplus = running[other][1] - sheet["L"][index].value
            goal -= possible_surplus
            calculateAndFill(sheet, CPD, goal, day, days, days_in_current_month, index + platform)
        else: calculateAndFill(sheet, CPD, goal, day, days, days_in_current_month, index + platform)
    elif platforms: # if one platform is paused, other platform performance is calculated with goal reduced/increased by possible surplus/deficit of clicks from paused platform, if any
        try: possible_surplus = (sheet["Q"][index].value - running[platform][1]) - sheet["L"][index].value
        except TypeError: possible_surplus = sheet["Q"][index].value - running[platform][1] # assumes 0 as goal if no goal specified
        goal -= possible_surplus
        calculateAndFill(sheet, CPD, goal, day, days, days_in_current_month, index + platform)
    else: # only one platform is present
        calculateAndFill(sheet, CPD, goal, day, days, days_in_current_month, index + platform)

def getCumulativeClicks(workbook, months_map, company, year, month):
    """
    Recursively fetches and sums clicks for every month of active status of campaign(s).
    Requires accurate data on past clicks in Prej column across all months and continuous active status of campaign(s).
    Returns cumulative number of clicks for company not including value for current month.
    :param workbook: openpyxl object; active workbook to work on
    :param months_map: dict; mapping of month keys (int) to spreadsheet values (str)
    :param year: int; starting year value
    :param month: int; starting month value
    :return: int; cumulative number of clicks for company across previous months
    """
    try: sheet = workbook[months_map[month] + " " + str(year)[2:]] # creates current spreadsheet
    except KeyError: return 0 # returns 0 if workbook runs out of spreadsheets
    for i in range(1, len(sheet["A"]), 2):
        if sheet["A"][i].value == company:
            before_GAdW = sheet["K"][i].value
            before_TSmedia = sheet["K"][i+1].value
            if before_GAdW == None or before_GAdW == "-": before_GAdW = 0
            if before_TSmedia == None or before_TSmedia == "-": before_TSmedia = 0
            month -= 1
            if month < 1:
                month = 12
                year -= 1
            return before_GAdW + before_TSmedia + getCumulativeClicks(workbook, months_map, company, year, month)
    return 0 # returns 0 if company not in current spreadsheet

def cumulativePercentageAndFill(spreadsheet, index, red, orange, yellow, blue_green, light_green):
    """
    Sets cumulative percentage value and applies fill sccording to clicks percentage.
    :param spreadsheet: openpyxl object; active XLS[X] spreadsheet to work on
    :param index: int; index of spreadsheet row to work on
    :param red: PatternFill object; fill for color red
    :param orange: PatternFill object; fill for color orange
    :param yellow: PatternFill object; fill for color yellow
    :param blue_green: PatternFill object; fill for color blue-green
    :param light_green: PatternFill object; fill for color light-green
    """
    spreadsheet["J" + str(index + 1)] = spreadsheet["I"][index].value / spreadsheet["I"][index + 1].value  # calculates cumulative % for Paket 49, 99, 199, 399
    if spreadsheet["J"][index].value < 0.25: spreadsheet["J" + str(index + 1)].fill = red  # applies fill according to cumulative clicks percentage
    elif spreadsheet["J"][index].value < 0.5: spreadsheet["J" + str(index + 1)].fill = orange
    elif spreadsheet["J"][index].value < 0.75: spreadsheet["J" + str(index + 1)].fill = yellow
    elif spreadsheet["J"][index].value < 1: spreadsheet["J" + str(index + 1)].fill = blue_green
    elif spreadsheet["J"][index].value == 1: spreadsheet["J" + str(index + 1)].fill = light_green
    else: spreadsheet["J" + str(index + 1)].fill = dark_green

def applyBorders(spreadsheet, top_border, bottom_border, top_left_border, top_right_border, bottom_left_border, bottom_right_border):
    """
    (Re)applies borders to cells in a spreadsheet.
    :param sheet: openpyxl object; active XLS[X] spreadsheet to work on
    """
    for i in range(1, len(spreadsheet["A"]), 2):
        if spreadsheet["A"][i].value:
            for column in ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O", "P", "Q"]:
                spreadsheet[column + str(i + 1)].border = top_border
                spreadsheet[column + str(i + 2)].border = bottom_border
            if month in [1, 3, 5, 7, 8, 10, 12]:
                for j in range(1, 32):
                    sheet[days[j][0] + str(j + 1)].border = top_left_border
                    sheet[days[j][1] + str(j + 1)].border = top_right_border
                    sheet[days[j][0] + str(j + 2)].border = bottom_left_border
                    sheet[days[j][1] + str(j + 2)].border = bottom_right_border
        else:
            for column in ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O", "P", "Q", "R"]:
                spreadsheet[column + str(i + 1)].border = top_border
                spreadsheet[column + str(i + 2)].border = bottom_border
            break

# creates path & backup
path = getcwd().replace("\\", "\\\\") + "\\\\"
copy2(path + "Poslovni-pregled.xlsx", path + "Poslovni-pregled_backup.xlsx")
print("Backup file created: {}Poslovni-pregled_backup.xlsx".format(path.replace("\\\\", "\\")))
# date variables
months = { 1: "JAN",
           2: "FEB",
           3: "MAR",
           4: "APR",
           5: "MAJ",
           6: "JUN",
           7: "JUL",
           8: "AVG",
           9: "SEP",
          10: "OKT",
          11: "NOV",
          12: "DEC"}
days =   { 1: ("S", "T"),
           2: ("U", "V"),
           3: ("W", "X"),
           4: ("Y", "Z"),
           5: ("AA", "AB"),
           6: ("AC", "AD"),
           7: ("AE", "AF"),
           8: ("AG", "AH"),
           9: ("AI", "AJ"),
          10: ("AK", "AL"),
          11: ("AM", "AN"),
          12: ("AO", "AP"),
          13: ("AQ", "AR"),
          14: ("AS", "AT"),
          15: ("AU", "AV"),
          16: ("AW", "AX"),
          17: ("AY", "AZ"),
          18: ("BA", "BB"),
          19: ("BC", "BD"),
          20: ("BE", "BF"),
          21: ("BG", "BH"),
          22: ("BI", "BJ"),
          23: ("BK", "BL"),
          24: ("BM", "BN"),
          25: ("BO", "BP"),
          26: ("BQ", "BR"),
          27: ("BS", "BT"),
          28: ("BU", "BV"),
          29: ("BW", "BX"),
          30: ("BY", "BZ"),
          31: ("CA", "CB")}
today = date.today()
year = str(today.year)[2:]
month = months[today.month]
day = today.day
prev_month_year = today.year
prev_month_month = today.month - 1
if prev_month_month == 0: # decrements year and sets previous month to 12 (December)
    prev_month_year = prev_month_year - 1
    prev_month_month = 12
if prev_month_month < 10: # stringifies previous year and month, adds 0 in front of month if month == single digit
    prev_month_year = str(prev_month_year)
    prev_month_month = "0" + str(prev_month_month)
else:
    prev_month_year = str(prev_month_year)
    prev_month_month = str(prev_month_month)
# importing clicks data from GAdW
try: imported = json.loads(sub(r"\s{2,}", " ", sub(r"\n", " ", open("GAdW_JSON.txt", encoding="UTF-8").read())))
except ValueError: imported = json.loads(sub(r"\s{2,}", " ", sub(r"\n", " ", open("GAdW_JSON.txt", encoding="UTF-8-SIG").read()))) # handles possible ValueError due to manual TXT file editing
except FileNotFoundError: imported = {"stats": {}, "date": [1000,1,1]} # creates meaningless object if file not found
stats_GAdW = imported["stats"]
date_GAdW = imported["date"]
year_GAdW = str(date_GAdW[0])[2:]
month_GAdW = months[date_GAdW[1]]
day_GAdW = date_GAdW[2]
if (year_GAdW, month_GAdW, day_GAdW) != (year, month, day): # creates empty dict if date from GAdW_JSON.txt does not match today
    stats_GAdW = {}
    print("No GAdW_JSON.txt for {} found.".format(today))
# importing clicks data from TSmedia (Graphite)
url_login = "http://graphite.tsmedia.si/dashboard/login"
url_data = "https://graphite.tsmedia.si/dashboard/campaigns-all"
payload = {"email": "damjan.mihelic@tsmedia.si", "password": "L!4)ggdSkRH-N/Sqe9Qq-/379YbO4X"}
stats_TSmedia = {}
with requests.Session() as s:
    login = s.post(url_login, data=payload)
    request = s.get(url_data)
    soup = bs(request.content, "html5lib")
campaigns_TSmedia = soup.find_all("tr")
for campaign in campaigns_TSmedia:
    try: # IndexError skips first element in list, which is an empty list
        if campaign.find_all("td", class_="right")[0].getText() == "✔": # gets only active campaigns (not ✘)
            OP = [element.getText() for element in campaign.find_all("div", class_="xbold")][0][:9]
            click_stats = [element.getText() for element in campaign.find_all("small", class_="xbold")]
            clicks = int(sub(r"\s.+", "", click_stats[2])) + int(sub(r"\s.+", "", click_stats[0])) # sums 'Monthly' and 'Today' categories
            stats_TSmedia[OP] = clicks
            impressions_CTR_stats = campaign.find_all("td", class_="right medium")
            impressions = int(impressions_CTR_stats[0].getText().replace(",", ""))
            CTR = float(impressions_CTR_stats[1].getText()[:-1])
            stats_TSmedia[OP] = [clicks, impressions, CTR]
    except IndexError: continue
# working on XLS file
wb = load_workbook("Poslovni-pregled.xlsx", data_only = True)
counter = 3 # cell counter to be incremented for applying borders and fills to total cost/percentage cells
try:
    sheet = wb[month + " " + year]
    companies = sheet["A"]
except KeyError: # create sheet for new month if not existing
    decrement_month = 1
    decrement_year = 1
    while True: # cycles back by month until last valid sheet is found
        last_found_month = today.month - decrement_month
        last_found_year = today.year
        if last_found_month == 0:
            last_found_month = 12
            last_found_year = last_found_year - decrement_year
        try:
            last_found_sheet = wb[months[last_found_month] + " " + str(last_found_year)[2:]]
            sheet = wb.copy_worksheet(last_found_sheet)
            sheet.title = month + " " + year
            companies = sheet["A"]
            break
        except KeyError:
            decrement_month += 1
            decrement_year = decrement_month // 12 + 1
top_border          = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'))
bottom_border       = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             bottom=Side(style='thin'))
top_left_border     = Border(left=Side(style='thin'),
                             top=Side(style='thin'))
top_right_border    = Border(right=Side(style='thin'),
                             top=Side(style='thin'))
bottom_left_border  = Border(left=Side(style='thin'),
                             bottom=Side(style='thin'))
bottom_right_border = Border(right=Side(style='thin'),
                             bottom=Side(style='thin'))
grey         = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
light_green  = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
blue_green   = PatternFill(start_color="63F292", end_color="63F292", fill_type="solid")
dark_green   = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
yellow       = PatternFill(start_color="FFDB69", end_color="FFDB69", fill_type="solid")
orange       = PatternFill(start_color="F79646", end_color="F79646", fill_type="solid")
red          = PatternFill(start_color="FF4F4F", end_color="FF4F4F", fill_type="solid")
no_fill      = PatternFill(fill_type=None)
days_in_current_month = max(max(Calendar().monthdayscalendar(today.year, today.month)))
for i in range(1, len(companies), 2): # applies gray fill for all days if only TSmedia or only GAdW, unless values found
    if sheet["A"][i].value:
        counter += 2
        platforms = (sheet["D"][i].value, sheet["D"][i+1].value)
        apply_grey = list() # assumes empty or single entry (at least one network always active)
        if "GAdW" not in platforms: apply_grey.append(i)
        if "TSmedia" not in platforms: apply_grey.append(i + 1)
        if apply_grey:
            for j in range(1, days_in_current_month + 1):
                if not sheet[days[j][0]][apply_grey[0]].value:
                    sheet[days[j][0]][apply_grey[0]].fill = grey
                    sheet[days[j][1]][apply_grey[0]].fill = grey
# clicks from previous month on GAdW
try:
    imported_prev = json.loads(sub(r"\s{2,}", " ", sub(r"\n", " ", open("GAdW_JSON_{}-{}.txt".format(prev_month_year, prev_month_month), encoding="UTF-8").read())))
    stats_prev = imported_prev["stats"]
    for i in range(1, len(sheet["A"]), 2): # clears values from columns "G" through "Q" and removes fill iff GAdW_JSON_YYYY-MM.txt found
        if sheet["A"][i].value:
            for column in ["G", "H", "J"]:
                sheet[column + str(i + 1)].value = None
                sheet[column + str(i + 1)].fill = grey
            for column in [ "I", "K", "L", "M", "N", "O", "P", "Q"]:
                sheet[column + str(i + 1)].value = None
                sheet[column + str(i + 2)].value = None
                sheet[column + str(i + 1)].fill = grey
                sheet[column + str(i + 2)].fill = grey
    for i in range(1, len(companies), 2): # starts with 2nd row to omit 1st row and jumps by 2 since each company takes 2 rows
        if companies[i].value:
            try:
                if sheet["K" + str(i + 1)].value in [None, 0]: sheet["K" + str(i + 1)] = stats_prev[sheet["C"][i].value][0] # ignores if values are already set
            except KeyError: print("{} / {} not found in GAdW stats.".format(companies[i].value, sheet["C"][i].value))
except FileNotFoundError: print("GAdW_JSON_{}-{}.txt".format(prev_month_year, prev_month_month) + " not found.")
# clicks from previous month on TSmedia
try:
    with open("campaign-ctr-{}-{}.csv".format(prev_month_year, prev_month_month), encoding="UTF-8") as TSmedia_prev:
        campaigns = list(reader(TSmedia_prev, delimiter=";"))
    for i in range(1, len(campaigns)):
        campaign = campaigns[i]
        if campaign[0][:2] == "OP":
            OP = campaign[0][:9]
            for i in range(1, len(sheet["C"]), 2):
                if sheet["C"][i].value == OP and sheet["K" + str(i + 2)].value in [None, 0]: sheet["K" + str(i + 2)] = int(campaign[4]) # ignores if values are already set
except FileNotFoundError: print("campaign-ctr-{}-{}.csv not found.".format(prev_month_year, prev_month_month))
# calculating new monthly goals & updating cumulative clicks
try:
    exceptions_prev = json.loads(open("Exceptions_prev_{}-{}.txt".format(prev_month_year, prev_month_month), encoding="UTF-8").read()) # adds past exceptions (additional clicks for TSmedia campaigns from other sources)
    print("Exceptions for previous month found and used!")
except FileNotFoundError:
    exceptions_prev = {}
    print("Exceptions_prev_JSON.txt not found.") # end of exceptions
goals = {49: (40, 160), 99: (80, 320), 199: (160, 640), 399: (320, 1280)}
sheet_prev_month = wb[months[int(prev_month_month)] + " " + prev_month_year[2:]]
goals_prev_month = dict()
for i in range(1, len(sheet_prev_month["L"]), 2):
    if sheet_prev_month["A"][i].value:
        goals_prev_month[sheet_prev_month["C"][i].value] = sheet_prev_month["L"][i].value + sheet_prev_month["L"][i+1].value
for i in range(1, len(sheet["A"]), 2):
    if sheet["A"][i].value and not sheet["L"][i].value and not sheet["L"][i+1].value: # ignores if goals fields have values (assuming goals already set)
        sheet["K" + str(i + 2)] = sheet["K"][i+1].value + exceptions_prev.get(sheet["C"][i].value, 0) # adds possible exceptions value to TSmedia previous month value
        GAdW_clicks = sheet["K"][i].value
        TSmedia_clicks = sheet["K"][i+1].value
        if not GAdW_clicks: GAdW_clicks = 0 # if value of clicks is NoneType
        if not TSmedia_clicks: TSmedia_clicks = 0 # if value of clicks is NoneType
        if sheet["E"][i].value not in goals.keys(): # if Paket is custom (not 49, 99, 199, 399), TSmedia only assumed
            try: new_goal = (0, goals_prev_month[sheet["C"][i].value] - (GAdW_clicks + TSmedia_clicks))
            except KeyError: new_goal = (0, sheet["E"][i].value)
        else: # regular Paket values (49, 99, 199, 399)
            surplus = GAdW_clicks + TSmedia_clicks - goals_prev_month.get(sheet["C"][i].value, 0) # can be negative (= deficit)
            packet = sheet["E"][i].value
            platforms = (sheet["D"][i].value, sheet["D"][i+1].value)
            try:
                if None in platforms and "TSmedia" in platforms: # if regular Paket contains only TSmedia network
                    new_goal = (0, sum(goals[packet]) - surplus)
                    # if new_goal[1] < 0: new_goal = (0, 0) # sets goal to 0 if surplus > total monthly goal
                elif None in platforms and "GAdW" in platforms: # if regular Paket contains only GAdW network
                    new_goal = (sum(goals[packet]) - surplus, 0)
                    # if new_goal[0] < 0: new_goal = (0, 0) # sets goal to 0 if surplus > total monthly goal
                else:
                    new_goal = (goals[packet][0], goals[packet][1] - surplus)
                    if new_goal[1] < 0: # if surplus > entire monthly goal -> negative TSmedia value is added to GAdW value (reducing GAdW value), while TSmedia goal is set to 0
                        negative_TSmedia = new_goal[1]
                        new_goal = (goals[packet][0] + negative_TSmedia, 0)
                        sheet["L"][i + 1].fill = dark_green # applies color fill as if TSmedia goal met
                        if new_goal[0] < 0: # if GAdW value is negative after previous step, GAdW goal is set to 0, while negative GAdW value is set as TSmedia goal
                            negative_GAdW = new_goal[0]
                            new_goal = (0, negative_GAdW)
                            sheet["L"][i].fill = dark_green # applies color fill as if GAdW goal met
                            sheet["M"][i].fill = dark_green # applies color fill as if total goal met
            except KeyError: new_goal = (0, goals_prev_month.get(sheet_prev_month["A"][i].value, packet) - surplus)
        #if sheet["L" + str(i + 1)].value != None or sheet["L" + str(i + 2)].value != None or sheet["M" + str(i + 1)].value != None: # updates new goal iff value not already set (does not overwrite manually entered values)
        sheet["L" + str(i + 1)] = new_goal[0]
        sheet["L" + str(i + 2)] = new_goal[1]
        sheet["M" + str(i + 1)] = new_goal[0] + new_goal[1]
    if sheet["E"][i].value in goals.keys() and sheet["A"][i].value and sheet["I"][i].value == None: # updates cumulative clicks for regular Paket iff cumulative clicks value not found
        GAdW_clicks = sheet["K"][i].value
        TSmedia_clicks = sheet["K"][i+1].value
        if not GAdW_clicks: GAdW_clicks = 0 # if value of clicks is NoneType
        if not TSmedia_clicks: TSmedia_clicks = 0 # if value of clicks is NoneType
        cumulative = GAdW_clicks + TSmedia_clicks
        prev_month_cumulative = 0
        prev_month_total = 0
        for j in range(1, len(sheet_prev_month["A"]), 2):
            if sheet_prev_month["A"][j].value == sheet["A"][i].value: # gets previous month cumulative and total clicks values if company names match
                prev_month_cumulative = sheet_prev_month["I"][j].value
                prev_month_total = sheet_prev_month["I"][j+1].value
        try:
            sheet["I" + str(i + 1)] = prev_month_cumulative + cumulative # sets new cumulative value (previous month cumulative + current month cumulative)
            sheet["I" + str(i + 2)] = prev_month_total + sum(goals[sheet["E"][i].value]) # sets new total clicks value (previous month total + packet monthly goal)
            cumulativePercentageAndFill(sheet, i, red, orange, yellow, blue_green, light_green)
        except TypeError: pass # skips over companies with no cumulative clicks value
    elif sheet["A"][i].value and sheet["I"][i].value == None: # updates cumulative clicks for custom Paket iff cumulative clicks value not found
        sheet["I" + str(i + 1)] = getCumulativeClicks(wb, months, sheet["A"][i].value, today.year, today.month) # sets new cumulative value (sum of cumulative values across previous months)
        sheet["I" + str(i + 2)] = sheet["E"][i].value # sets total clicks value (same value across all months, defined as custom Paket value)
        cumulativePercentageAndFill(sheet, i, red, orange, yellow, blue_green, light_green)
applyBorders(sheet_prev_month, top_border, bottom_border, top_left_border, top_right_border, bottom_left_border, bottom_right_border)
# current clicks & impressions & CTR on GAdW
paused = mergePausedFiles(path)
if not paused: print("GAdW_JSON_paused.txt not found.")
for i in range(1, len(sheet["A"]), 2): # starts with 2nd row to omit 1st row and jumps by 2 since each company takes 2 rows
    if sheet["A"][i].value:
        if paused: # adds clicks and cost values for paused campaigns to day cell corresponding to date of pausing
            for paused_campaign in paused:
                try:
                    paused_values = paused_campaign["stats"][sheet["C"][i].value]
                    try: del stats_GAdW[sheet["C"][i].value] # removes account found in paused values from GAdW stats (if present) to prevent double processing
                    except KeyError: pass
                    sheet[days[paused_campaign["date"][2]][0] + str(i + 1)] = paused_values[0] # updates clicks
                    if paused_values[0] == sheet["L"][i].value: sheet["L"][i].fill = light_green # applies color fill according to clicks/goals
                    elif paused_values[0] > sheet["L"][i].value: sheet["L"][i].fill = dark_green
                    elif paused_values[0] >= round(sheet["L"][i].value / 2): sheet["L"][i].fill = yellow
                    else: sheet["L"][i].fill = red
                    sheet["G" + str(i + 1)] = paused_values[1] # updates cost
                    sheet["H" + str(i + 1)] = sheet["G"][i].value / sheet["F"][i].value
                    try:
                        if sheet["G"][i].value / sheet["F"][i].value > 1: sheet["H"][i].fill = red
                        elif sheet["G"][i].value / sheet["F"][i].value >= 0.5: sheet["H"][i].fill = yellow
                        else: sheet["H"][i].fill = light_green
                    except ValueError: pass
                    pauseToEndOfMonth(sheet, paused_campaign["date"][2], days, days_in_current_month, grey, "GAdW")
                except KeyError: pass
        try:
            value = stats_GAdW[sheet["C"][i].value]
            if value[0] == "paused": pauseToEndOfMonth(sheet, day, days, days_in_current_month, grey, "GAdW") # marks paused campaigns until end of month
            else:
                if not sheet[days[day][0]][i].value or value[0] > sheet[days[day][0]][i].value: sheet[days[day][0] + str(i + 1)] = value[0] # updates clicks only if old value is lower or None
                unpauseToEndOfMonth(sheet, day, days, days_in_current_month, no_fill, "GAdW") # marks cells possibly marked as paused as unpaused
                if value[0] == sheet["L"][i].value: sheet["L"][i].fill = light_green # applies color fill according to clicks/goals
                elif value[0] > sheet["L"][i].value: sheet["L"][i].fill = dark_green
                elif value[0] >= round(sheet["L"][i].value * 0.75): sheet["L"][i].fill = blue_green
                elif value[0] >= round(sheet["L"][i].value * 0.5): sheet["L"][i].fill = yellow
                elif value[0] >= round(sheet["L"][i].value * 0.25): sheet["L"][i].fill = orange
                else: sheet["L"][i].fill = red
                if not sheet["G"][i].value or value[2] > sheet["G"][i].value:
                    sheet["G" + str(i + 1)] = value[2] # updates cost only if old value is lower or None
                sheet["H" + str(i + 1)] = sheet["G"][i].value / sheet["F"][i].value
                try:
                    if sheet["G"][i].value / sheet["F"][i].value > 1: sheet["H"][i].fill = red
                    elif sheet["G"][i].value / sheet["F"][i].value >= 0.75: sheet["H"][i].fill = orange
                    elif sheet["G"][i].value / sheet["F"][i].value >= 0.5: sheet["H"][i].fill = yellow
                    elif sheet["G"][i].value / sheet["F"][i].value >= 0.25: sheet["H"][i].fill = blue_green
                    else: sheet["H"][i].fill = light_green
                except ValueError: pass
                sheet["N" + str(i + 1)] = value[3] # updates impressions
                sheet["O" + str(i + 1)] = value[1] # updates CTR
        except KeyError: print("{} / {} not found in GAdW stats.".format(companies[i].value, sheet["C"][i].value))
# current clicks & impressions & CTR on TSmedia
try:
    exceptions = json.loads(open("Exceptions_JSON.txt", encoding="UTF-8").read()) # adds current exceptions (additional clicks for TSmedia campaigns from other sources)
    print("Exceptions for current month found and used!")
except FileNotFoundError:
    exceptions = None
    print("Exceptions_JSON.txt not found.") # end of exceptions
for i in range(1, len(sheet["A"]), 2):
    if sheet["A"][i].value:
        try:
            new_clicks = stats_TSmedia.get(sheet["C"][i].value, None)[0]
            new_impressions = stats_TSmedia.get(sheet["C"][i].value, None)[1]
            new_CTR = stats_TSmedia.get(sheet["C"][i].value, None)[2] / 100
        except TypeError:
            new_clicks = None
            new_impressions = None
            new_CTR = None
        if new_clicks != None and new_impressions > 0:
            unpauseToEndOfMonth(sheet, day, days, days_in_current_month, no_fill, "TSmedia")  # marks cells possibly marked as paused as unpaused
            sheet[days[day][0] + str(i + 2)] = new_clicks
        if exceptions:
            OP = sheet["C"][i].value
            if OP in exceptions.keys() and sheet[days[day][0]][i+1].value: sheet[days[day][0] + str(i + 2)] = \
            sheet[days[day][0]][i+1].value + exceptions[OP]
        clicks_TSmedia = sheet[days[day][0]][i+1].value
        if clicks_TSmedia != None: # applies color fill according to clicks/goals
            if clicks_TSmedia == sheet["L"][i+1].value: sheet["L"][i+1].fill = light_green
            elif clicks_TSmedia > sheet["L"][i+1].value: sheet["L"][i+1].fill = dark_green
            elif clicks_TSmedia >= round(sheet["L"][i+1].value * 0.75): sheet["L"][i+1].fill = blue_green
            elif clicks_TSmedia >= round(sheet["L"][i+1].value * 0.5): sheet["L"][i+1].fill = yellow
            elif clicks_TSmedia >= round(sheet["L"][i+1].value * 0.25): sheet["L"][i+1].fill = yellow
            else: sheet["L"][i+1].fill = red
        else: # marks paused campaigns until end of month
            try:
                if new_impressions > 0: pauseToEndOfMonth(sheet, day, days, days_in_current_month, grey, "TSmedia")
            except TypeError: pauseToEndOfMonth(sheet, day, days, days_in_current_month, grey, "TSmedia")
        if new_impressions != None: sheet["N" + str(i + 2)] = new_impressions # updates impressions
        if new_CTR != None: sheet["O" + str(i + 2)] = new_CTR # updates CTR
# current total clicks & impressions % CTR & performance
for i in range(1, len(sheet["A"]), 2):
    if sheet["A"][i].value:
        clicks_GAdW = sheet[days[day][0]][i].value
        clicks_TSmedia = sheet[days[day][0]][i+1].value
        if clicks_GAdW or clicks_TSmedia: # checks if value for clicks is None due to paused campaign
            if not clicks_GAdW: clicks_GAdW = findLastValue(sheet, i, days, day)
            if not clicks_TSmedia: clicks_TSmedia = findLastValue(sheet, i + 1, days, day)
            total = clicks_GAdW + clicks_TSmedia
            total_goal = sheet["M"][i].value
            if total_goal == None: # if total goal value was omitted when adding new entry to the table
                sheet["M" + str(i + 1)] = sheet["L"][i].value + sheet["L"][i+1].value
                total_goal = sheet["L"][i].value + sheet["L"][i+1].value
            if total_goal <= 0: # if surplus >= total monthly goal
                sheet["Q" + str(i + 1)] = total
                sheet["Q" + str(i + 2)] = "N/A" # sets percentage of completed goal to N/A
                sheet["M"][i].fill = dark_green # applies color fill as if goals surpassed
                sheet["Q"][i].fill = dark_green
                sheet["Q"][i+1].fill = dark_green
            else:
                sheet["Q" + str(i + 1)] = total
                sheet["Q" + str(i + 2)] = total / total_goal
                if total == sheet["M"][i].value: # applies color fill according to total clicks/goals
                    sheet["M"][i].fill = light_green
                    sheet["Q"][i].fill = light_green
                    sheet["Q"][i+1].fill = light_green
                elif total > sheet["M"][i].value:
                    sheet["M"][i].fill = dark_green
                    sheet["Q"][i].fill = dark_green
                    sheet["Q"][i+1].fill = dark_green
                elif total >= round(sheet["M"][i].value * 0.75):
                    sheet["M"][i].fill = blue_green
                    sheet["Q"][i].fill = blue_green
                    sheet["Q"][i+1].fill = blue_green
                elif total >= round(sheet["M"][i].value * 0.5):
                    sheet["M"][i].fill = yellow
                    sheet["Q"][i].fill = yellow
                    sheet["Q"][i+1].fill = yellow
                elif total >= round(sheet["M"][i].value * 0.25):
                    sheet["M"][i].fill = orange
                    sheet["Q"][i].fill = orange
                    sheet["Q"][i+1].fill = orange
                else:
                    sheet["M"][i].fill = red
                    sheet["Q"][i].fill = red
                    sheet["Q"][i+1].fill = red
            running = [(sheet["D"][i].value, sheet[days[day][0]][i].value), (sheet["D"][i+1].value, sheet[days[day][0]][i+1].value)]
            if running[0][0] and running[1][0]: # checks if both platforms are present
                if running[0][1]: # applies performance % iff value for clicks found in cell
                    CPD_GAdW = clicks_GAdW / day  # clicks per day on GAdW
                    goal_GAdW = sheet["L"][i].value
                    performance(running, sheet, CPD_GAdW, goal_GAdW, day, days, days_in_current_month, i, 0)
                if running[1][1]: # applies performance % iff value for clicks found in cell
                    CPD_TSmedia = clicks_TSmedia / day # clicks per day on TSmedia
                    goal_TSmedia = sheet["L"][i+1].value
                    performance(running, sheet, CPD_TSmedia, goal_TSmedia, day, days, days_in_current_month, i, 1)
            elif running[0][0] == "GAdW": # only GAdW platform present
                if running[0][1]: # applies performance % iff value for clicks found in cell
                    CPD_GAdW = clicks_GAdW / day # clicks per day on GAdW
                    goal_GAdW = sheet["L"][i].value
                    performance(running, sheet, CPD_GAdW, goal_GAdW, day, days, days_in_current_month, i, 0, platforms=False)
            else: # only TSmedia platform present
                if running[1][1]: # applies performance % iff value for clicks found in cell
                    CPD_TSmedia = clicks_TSmedia / day # clicks per day on TSmedia
                    goal_TSmedia = sheet["L"][i+1].value
                    performance(running, sheet, CPD_TSmedia, goal_TSmedia, day, days, days_in_current_month, i, 1, platforms=False)
        try:
            sheet["P" + str(i + 1)] = sheet["N"][i].value + sheet["N"][i+1].value # total impressions
        except TypeError:
            if sheet["N"][i].value == None and sheet["N"][i+1].value == None: sheet["P" + str(i + 1)] = 0
            elif sheet["N"][i].value == None: sheet["P" + str(i + 1)] = sheet["N"][i+1].value
            else: sheet["P" + str(i + 1)] = sheet["N"][i].value
        try:
            sheet["P" + str(i + 2)] = sheet["Q"][i].value / sheet["P"][i].value # total CTR
        except (TypeError, ZeroDivisionError) as e: pass
# cumulative clicks for custom Paket campaigns (requires current clicks to be updated first)
for i in range(1, len(sheet["A"]), 2):
    if sheet["A"][i].value and sheet["I"][i].value != None: # updates cumulative clicks only if cumulative clicks value found
        if sheet["E"][i].value not in goals.keys() and sheet["I"][i+1].value != 0: # calculates cumulative % for custom Paket by adding current clicks to numerator
            try:
                sheet["I" + str(i + 1)] = getCumulativeClicks(wb, months, sheet["A"][i].value, today.year, today.month) + sheet["N"][i].value
                sheet["J" + str(i + 1)] = sheet["I"][i].value / sheet["I"][i+1].value
                if sheet["J"][i].value < 0.25: sheet["J" + str(i + 1)].fill = red # applies fill according to cumulative clicks percentage
                elif sheet["J"][i].value < 0.5: sheet["J" + str(i + 1)].fill = orange
                elif sheet["J"][i].value < 0.75: sheet["J" + str(i + 1)].fill = yellow
                elif sheet["J"][i].value < 1: sheet["J" + str(i + 1)].fill = blue_green
                elif sheet["J"][i].value == 1: sheet["J" + str(i + 1)].fill = light_green
                else: sheet["J" + str(i + 1)].fill = dark_green
            except TypeError: pass # ignores if custom campaign is not yet active
# formatting borders
applyBorders(sheet, top_border, bottom_border, top_left_border, top_right_border, bottom_left_border, bottom_right_border)
# current total costs
budget = []
costs = []
for i in range(1, len(sheet["A"]), 2):
    if sheet["A"][i].value:
        if type(sheet["F"][i].value) in [int, float]: budget.append(sheet["F"][i].value)
        if type(sheet["G"][i].value) in [int, float]: costs.append(sheet["G"][i].value)
sheet["F" + str(counter + 1)] = sum(budget)
sheet["G" + str(counter + 1)] = sum(costs)
sheet["H" + str(counter + 1)] = sum(costs) / sum(budget)
if sheet["H" + str(counter + 1)].value > 1:
    sheet["H" + str(counter + 1)].fill = red
    sheet["H" + str(counter + 2)].fill = red
elif sheet["H" + str(counter + 1)].value >= 0.75:
    sheet["H" + str(counter + 1)].fill = orange
    sheet["H" + str(counter + 2)].fill = orange
elif sheet["H" + str(counter + 1)].value >= 0.5:
    sheet["H" + str(counter + 1)].fill = yellow
    sheet["H" + str(counter + 2)].fill = yellow
elif sheet["H" + str(counter + 1)].value >= 0.25:
    sheet["H" + str(counter + 1)].fill = blue_green
    sheet["H" + str(counter + 2)].fill = blue_green
else:
    sheet["H" + str(counter + 1)].fill = light_green
    sheet["H" + str(counter + 2)].fill = light_green
sheet["F" + str(counter + 1)].border = top_border
sheet["F" + str(counter + 2)].border = bottom_border
sheet["G" + str(counter + 1)].border = top_border
sheet["G" + str(counter + 2)].border = bottom_border
sheet["H" + str(counter + 1)].border = top_border
sheet["H" + str(counter + 2)].border = bottom_border
# writing update info
sheet["A" + str(counter + 4)] = "Last updated: " + datetime.now().strftime("%Y-%m-%d %H:%M:%S")
# saving file
while True:
    try:
        wb.save(path + "Poslovni-pregled.xlsx")
        break
    except PermissionError: input("\nPlease close Poslovni-pregled.xlsx and press any key. ")
print("\nArticle info saved as Poslovni-pregled.xlsx")